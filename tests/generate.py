"""Génère les datasets et configs pour les 6 scénarios de test.

Usage : python generate.py

Chaque scénario est dans son dossier tests/scenario_NN/.
"""
import openpyxl
import secrets
import shutil
from pathlib import Path

rng = secrets.SystemRandom()
BASE = Path(__file__).parent

# ── Scénario 01 : Standard (tout passe) ──────────

def gen_01():
    """5000 lignes, dataset équilibré, 3 règles."""
    d = BASE / "scenario_01"; d.mkdir(exist_ok=True)

    wb = openpyxl.Workbook(); ws = wb.active
    ws.cell(row=1, column=1, value="ID")
    ws.cell(row=1, column=3, value="Taille")
    ws.cell(row=1, column=4, value="Email")
    ws.cell(row=1, column=6, value="Genre")
    ws.cell(row=1, column=20, value="Region")

    regions = ["PACA"]*1000 + ["NAQ"]*1000 + ["IDF"]*1000 + ["BRE"]*1000 + ["OCC"]*1000
    rng.shuffle(regions)
    genders = ["Femme"]*2500 + ["Homme"]*2500
    rng.shuffle(genders)
    heights = [190]*1500 + [170]*2000 + [150]*1500  # 30% >180
    rng.shuffle(heights)

    for i in range(5000):
        r = i + 2
        ws.cell(row=r, column=1, value=f"ID_{i:05d}")
        ws.cell(row=r, column=3, value=heights[i])
        ws.cell(row=r, column=4, value=f"user{i:05d}@test.com")
        ws.cell(row=r, column=6, value=genders[i])
        ws.cell(row=r, column=20, value=regions[i])
    wb.save(d / "data.xlsx")

    (d / "config.yaml").write_text("""input:
  file: data.xlsx
  sheet: 0
email_column: D
winners: 200
batch_size: 50
rules:
  - column: F
    op: "="
    value: "Femme"
    percent: 50
  - column: T
    op: "="
    value: "PACA"
    percent: 20
  - column: C
    op: ">"
    value: 180
    percent: 30
""")
    print(f"[OK] scenario_01 : 5000 lignes, équilibré")


# ── Scénario 02 : Repêchage ──────────────────────

def gen_02():
    """5000 lignes, seulement 2% PACA → règles impossibles."""
    d = BASE / "scenario_02"; d.mkdir(exist_ok=True)

    wb = openpyxl.Workbook(); ws = wb.active
    ws.cell(row=1, column=1, value="ID")
    ws.cell(row=1, column=3, value="Taille")
    ws.cell(row=1, column=4, value="Email")
    ws.cell(row=1, column=6, value="Genre")
    ws.cell(row=1, column=20, value="Region")

    regions = ["PACA"]*100 + ["IDF"]*4900  # 2% PACA
    rng.shuffle(regions)
    genders = ["Femme"]*2500 + ["Homme"]*2500
    rng.shuffle(genders)
    heights = [190]*1500 + [170]*2000 + [150]*1500
    rng.shuffle(heights)

    for i in range(5000):
        r = i + 2
        ws.cell(row=r, column=1, value=f"ID_{i:05d}")
        ws.cell(row=r, column=3, value=heights[i])
        ws.cell(row=r, column=4, value=f"user{i:05d}@test.com")
        ws.cell(row=r, column=6, value=genders[i])
        ws.cell(row=r, column=20, value=regions[i])
    wb.save(d / "data.xlsx")

    # Mêmes règles que 01
    shutil.copy(BASE / "scenario_01" / "config.yaml", d / "config.yaml")
    print(f"[OK] scenario_02 : 5000 lignes, 2% PACA → repêchage")


# ── Scénario 03 : Plus de gagnants que de lignes ──

def gen_03():
    """100 lignes, winners=500."""
    d = BASE / "scenario_03"; d.mkdir(exist_ok=True)

    wb = openpyxl.Workbook(); ws = wb.active
    ws.cell(row=1, column=1, value="ID")
    ws.cell(row=1, column=4, value="Email")

    for i in range(100):
        r = i + 2
        ws.cell(row=r, column=1, value=f"ID_{i:03d}")
        ws.cell(row=r, column=4, value=f"user{i:03d}@test.com")
    wb.save(d / "data.xlsx")

    (d / "config.yaml").write_text("""input:
  file: data.xlsx
  sheet: 0
email_column: D
winners: 500
batch_size: 50
rules:
  - column: A
    op: "!="
    value: "NEXISTE_PAS"
    percent: 50
""")
    print(f"[OK] scenario_03 : 100 lignes, 500 gagnants → toutes les lignes")


# ── Scénario 04 : Sans règles ────────────────────

def gen_04():
    """10 000 lignes, aucune règle."""
    d = BASE / "scenario_04"; d.mkdir(exist_ok=True)

    wb = openpyxl.Workbook(); ws = wb.active
    ws.cell(row=1, column=1, value="ID")
    ws.cell(row=1, column=4, value="Email")

    for i in range(10000):
        r = i + 2
        ws.cell(row=r, column=1, value=f"ID_{i:05d}")
        ws.cell(row=r, column=4, value=f"user{i:05d}@test.com")
    wb.save(d / "data.xlsx")

    (d / "config.yaml").write_text("""input:
  file: data.xlsx
  sheet: 0
email_column: D
winners: 50
batch_size: 25
rules: []
""")
    print(f"[OK] scenario_04 : 10 000 lignes, 0 règle")


# ── Scénario 05 : Opérateurs numériques ──────────

def gen_05():
    """5000 lignes, âges 18-80, règles sur colonne numérique."""
    d = BASE / "scenario_05"; d.mkdir(exist_ok=True)

    wb = openpyxl.Workbook(); ws = wb.active
    ws.cell(row=1, column=1, value="ID")
    ws.cell(row=1, column=3, value="Age")
    ws.cell(row=1, column=4, value="Email")

    ages = [rng.randint(18, 80) for _ in range(5000)]
    for i in range(5000):
        r = i + 2
        ws.cell(row=r, column=1, value=f"ID_{i:05d}")
        ws.cell(row=r, column=3, value=ages[i])
        ws.cell(row=r, column=4, value=f"user{i:05d}@test.com")
    wb.save(d / "data.xlsx")

    (d / "config.yaml").write_text("""input:
  file: data.xlsx
  sheet: 0
email_column: D
winners: 100
batch_size: 50
rules:
  - column: C
    op: "<"
    value: 25
    percent: 15
  - column: C
    op: ">"
    value: 50
    percent: 40
""")
    print(f"[OK] scenario_05 : 5000 lignes, âges 18-80, opérateurs < >")


# ── Scénario 06 : Stress (5 règles, 20k lignes) ──

def gen_06():
    """20 000 lignes, 5 règles emboîtées."""
    d = BASE / "scenario_06"; d.mkdir(exist_ok=True)

    wb = openpyxl.Workbook(); ws = wb.active
    ws.cell(row=1, column=1, value="ID")
    ws.cell(row=1, column=3, value="Score")
    ws.cell(row=1, column=4, value="Email")
    ws.cell(row=1, column=5, value="Niveau")
    ws.cell(row=1, column=6, value="Categorie")
    ws.cell(row=1, column=7, value="Statut")

    scores = [rng.randint(0, 100) for _ in range(20000)]
    niveaux = ["A"]*4000 + ["B"]*6000 + ["C"]*10000
    rng.shuffle(niveaux)
    categories = ["X"]*6000 + ["Y"]*8000 + ["Z"]*6000
    rng.shuffle(categories)
    statuts = ["Actif"]*12000 + ["Inactif"]*8000
    rng.shuffle(statuts)

    for i in range(20000):
        r = i + 2
        ws.cell(row=r, column=1, value=f"ID_{i:05d}")
        ws.cell(row=r, column=3, value=scores[i])
        ws.cell(row=r, column=4, value=f"user{i:05d}@test.com")
        ws.cell(row=r, column=5, value=niveaux[i])
        ws.cell(row=r, column=6, value=categories[i])
        ws.cell(row=r, column=7, value=statuts[i])
    wb.save(d / "data.xlsx")

    (d / "config.yaml").write_text("""input:
  file: data.xlsx
  sheet: 0
email_column: D
winners: 500
batch_size: 100
rules:
  - column: G
    op: "="
    value: "Actif"
    percent: 80
  - column: F
    op: "="
    value: "X"
    percent: 30
  - column: E
    op: "="
    value: "A"
    percent: 20
  - column: F
    op: "="
    value: "Y"
    percent: 40
  - column: C
    op: ">"
    value: 50
    percent: 50
""")
    print(f"[OK] scenario_06 : 20 000 lignes, 5 règles emboîtées")


# ── Main ─────────────────────────────────────────

if __name__ == "__main__":
    gen_01()
    gen_02()
    gen_03()
    gen_04()
    gen_05()
    gen_06()
    print(f"\n{6} scénarios générés dans {BASE}/")
