"""Lecture de fichier Excel — mode read_only, colonnes par lettre."""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def read(filepath: Path, sheet: str | int = 0) -> tuple[dict, list[dict]]:
    """Lit un fichier Excel.

    Retourne (col_map, rows) où :
      col_map : {lettre: nom_colonne}  ex: {"A": "Nom", "B": "Email"}
      rows    : [{lettre: valeur}, ...]
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[sheet]] if isinstance(sheet, int) else wb[sheet]

    # Lire toutes les lignes en une fois avec iter_rows (rapide en read_only)
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return {}, []

    # En-têtes (première ligne)
    headers = all_rows[0]
    col_map = {}
    for col_idx, raw in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        col_map[letter] = str(raw) if raw is not None else f"Col_{letter}"

    # Données (lignes 2+)
    rows = []
    for data in all_rows[1:]:
        row = {}
        for col_idx, value in enumerate(data, start=1):
            letter = get_column_letter(col_idx)
            row[letter] = value
        rows.append(row)

    return col_map, rows
